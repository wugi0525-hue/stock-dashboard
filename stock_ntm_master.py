import yfinance as yf
import pandas as pd
import FinanceDataReader as fdr
from datetime import datetime, timedelta, date
import os
import time
import requests
import re
from io import StringIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

# 구글 시트 연결을 위한 도구 추가
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ==========================================
# 1. 설정 및 매핑
# ==========================================
FILE_NAME = "Global_NTM_Master_DB.xlsx"
DUAL_CLASS_MAPPING = {
    'FOX': 'FOXA', 'NWS': 'NWSA', 'GOOG': 'GOOGL', 'UHAL': 'UHAL-B'
}


def make_df_gspread_safe(df):
    """구글 시트에 업로드하기 전에 datetime 계열을 문자열로 변환해 JSON 직렬화 오류를 방지합니다."""
    def convert_cell(x):
        if pd.isna(x):
            return ""
        if isinstance(x, (pd.Timestamp, datetime, date)):
            return x.strftime('%Y-%m-%d')
        return x

    # 입력이 DataFrame이 아닐 수 있으니 강제 변환
    if not isinstance(df, pd.DataFrame):
        try:
            df = pd.DataFrame(df)
        except Exception:
            return df

    # 일반적으로 applymap이 존재하지만, 환경에 따라 없을 수 있으니 예외 처리
    try:
        return df.applymap(convert_cell)
    except AttributeError:
        for col in df.columns:
            try:
                df[col] = df[col].apply(convert_cell)
            except Exception:
                df[col] = df[col].astype(str).fillna("")
        return df

def get_wiki_session():
    """ 위키피디아 차단 방지용 헤더 """
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })
    return session


def upload_excel_sheets_to_gspread(client, spreadsheet_name, file_name, sheet_names):
    """주어진 엑셀 파일의 여러 시트를 구글 스프레드시트의 각각의 워크시트로 업로드합니다.
    존재하지 않는 워크시트는 새로 생성합니다.
    """
    # 스프레드시트 열기
    spreadsheet = client.open(spreadsheet_name)

    results = {'uploaded': [], 'failed': []}

    for sheet in sheet_names:
        print(f"Uploading sheet: {sheet} ...")
        try:
            df = pd.read_excel(file_name, sheet_name=sheet)
        except Exception as e:
            print(f" - Sheet read skipped: {sheet} ({e})")
            results['failed'].append((sheet, f"read_error: {e}"))
            continue

        df = df.fillna("")
        df_safe = make_df_gspread_safe(df)

        try:
            created = False
            try:
                worksheet = spreadsheet.worksheet(sheet)
            except Exception:
                worksheet = spreadsheet.add_worksheet(title=sheet, rows=max(100, len(df_safe)+5), cols=max(10, len(df_safe.columns)+2))
                created = True

            worksheet.clear()
            worksheet.update([df_safe.columns.values.tolist()] + df_safe.values.tolist())

            if created:
                print(f" - Created and uploaded: {sheet}")
            else:
                print(f" - Updated existing worksheet: {sheet}")

            results['uploaded'].append(sheet)
        except Exception as e:
            print(f" - Upload failed for sheet {sheet}: {e}")
            import traceback; traceback.print_exc()
            results['failed'].append((sheet, str(e)))
            continue

    print(f"Upload summary: {len(results['uploaded'])} uploaded, {len(results['failed'])} failed")
    return results

# ==========================================
# 2. 리스트 수집 (정교한 방식)
# ==========================================
def get_sp500_tickers():
    print("🇺🇸 S&P 500 리스트 수집 중...", end=" ")
    try:
        url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
        session = get_wiki_session()
        response = session.get(url)
        table = pd.read_html(StringIO(response.text))
        df = table[0]
        tickers = df['Symbol'].tolist()
        tickers = [t.replace('.', '-') for t in tickers]
        print(f"✅ {len(tickers)}개")
        return tickers
    except Exception as e:
        print(f"❌ 실패: {e}")
        return []

def get_kospi200_cleaned_tickers():
    print("🇰🇷 KOSPI 200 (ETF 정밀 제거) 수집 중...", end=" ")
    try:
        df_krx = fdr.StockListing('KOSPI')
        # ETF/ETN 키워드로 정교하게 필터링
        etf_keywords = 'KODEX|TIGER|KBSTAR|KOSEF|SOL|HANARO|ACE|ARIRANG|ETF|ETN'
        df_companies = df_krx[~df_krx['Name'].str.contains(etf_keywords, regex=True, case=False)]
        
        # 시총 상위 200개
        sort_col = 'Marcap' if 'Marcap' in df_companies.columns else 'MarketCap'
        df_top200 = df_companies.sort_values(by=sort_col, ascending=False).head(200)
        
        tickers = df_top200['Code'].tolist()
        tickers = [f"{code}.KS" for code in tickers]
        print(f"✅ {len(tickers)}개")
        return tickers
    except Exception as e:
        print(f"❌ 실패: {e}")
        return []

# ==========================================
# 3. 지수 현재가 가져오기 (기능 부활)
# ==========================================
def get_index_price(ticker):
    """ Summary 시트에 넣을 지수(S&P 500, KOSPI 200) 현재가를 가져옵니다. """
    try:
        data = yf.Ticker(ticker)
        # 1차 시도: fast_info (빠름)
        price = data.fast_info.get('last_price')
        if not price:
            price = data.info.get('regularMarketPrice')
        
        # 2차 시도: history (확실함)
        if not price:
            hist = data.history(period="1d")
            if not hist.empty:
                price = hist['Close'].iloc[-1]
                
        return price if price else 0
    except:
        return 0

# ==========================================
# 4. 핵심 분석 엔진 (선생님 검증용 정밀 로직)
# ==========================================
def get_analyst_count_naver(code):
    """
    네이버 증권에서 해당 종목의 투자의견(리포트 수) 정보를 크롤링합니다.
    URL: https://finance.naver.com/item/main.naver?code={code}
    패턴: '투자의견' 섹션 근처의 (숫자) 형태를 찾습니다.
    """
    try:
        # 1) 우선 모바일 통합 API를 시도합니다 (더 구조적이며 신뢰도가 높음)
        api_url = f"https://m.stock.naver.com/api/stock/{code}/integration"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        resp = requests.get(api_url, headers=headers, timeout=8)
        if resp.status_code == 200:
            try:
                data = resp.json()
                # 'researches' 키가 있으면 리포트 목록으로 보통 참여 애널리스트/리포트 항목 수를 반환
                if 'researches' in data and isinstance(data['researches'], list):
                    return len(data['researches'])

                # 일부 응답은 'consensusInfo'만 포함합니다. 여기엔 참여 수가 없을 수 있으므로 제외
                # 다른 후보로 'totalInfos'나 기타 항목에서 획득 시도
                if 'totalInfos' in data:
                    try:
                        for info in data['totalInfos']:
                            # 일부 항목에 'key' 혹은 'code'로 리포트 수가 표기될 수 있음
                            if info.get('code') in ('cnsCount', 'researchCount'):
                                val = re.search(r"(\d+)", str(info.get('value')))
                                if val:
                                    return int(val.group(1))
                    except Exception:
                        pass
            except ValueError:
                pass

        # 2) 모바일 API 실패하거나 정보가 없으면 메인 페이지 HTML에서 보수적으로 추출
        page_url = f"https://finance.naver.com/item/main.naver?code={code}"
        page = requests.get(page_url, headers=headers, timeout=8)
        if page.status_code != 200:
            return 0

        # 여러 인코딩 시도 (안전성 확보)
        text = ''
        try:
            page.encoding = 'euc-kr'
            text = page.text
        except Exception:
            try:
                text = page.content.decode('utf-8', errors='ignore')
            except Exception:
                text = page.text

        # 안전하게 '투자의견' 주변에서 괄호 숫자 패턴을 찾음
        try:
            idx = text.find('투자의견')
            if idx != -1:
                sub = text[idx:idx+800]
            else:
                sub = text[:1000]

            m = re.search(r"(\d+\.\d+)\s*\((\d+)\)", sub)
            if m:
                return int(m.group(2))

            # 대안 패턴: 단순 '(23)' 형태로 존재할 수 있으므로 첫 번째 괄호 숫자 추출
            m2 = re.search(r"\((\d+)\)", sub)
            if m2:
                return int(m2.group(1))
        except Exception:
            pass

        return 0
    except Exception:
        return 0

def analyze_ntm_data_robust(tickers):
    today = datetime.now()
    today_str = today.strftime("%Y-%m-%d") # 엑셀 저장용 문자열
    results = []
    
    total_count = len(tickers)
    print(f"\n🚀 총 {total_count}개 기업 정밀 NTM 분석 시작\n")

    for idx, symbol in enumerate(tickers):
        if (idx + 1) % 50 == 0: print(f"Processing {idx+1}/{total_count}...")
            
        try:
            tk = yf.Ticker(symbol)
            
            # 1. Info & Market Cap
            try:
                info = tk.info
                current_price = info.get('currentPrice')
                sector = info.get('sector', 'Unknown')
                short_name = info.get('shortName', symbol)
                market_cap = info.get('marketCap') 
            except:
                continue

            if not current_price: continue

            # [Analyst Count] 모든 종목에 대해 yfinance의 값을 우선 사용
            analyst_count = info.get('numberOfAnalystOpinions', 0)
            if analyst_count is None:
                analyst_count = 0

            # 2. 회계연도 계산
            fye_timestamp = info.get('nextFiscalYearEnd')
            if fye_timestamp:
                next_fye_date = datetime.fromtimestamp(fye_timestamp)
            else:
                next_fye_date = datetime(today.year, 12, 31)

            if next_fye_date < today:
                next_fye_date = next_fye_date.replace(year=next_fye_date.year + 1)

            # 3. EPS Estimate (Dual Class 매핑 로직 포함)
            estimates = tk.earnings_estimate
            
            if (estimates is None or estimates.empty) and (symbol in DUAL_CLASS_MAPPING):
                target_brother = DUAL_CLASS_MAPPING[symbol]
                try:
                    brother_tk = yf.Ticker(target_brother)
                    estimates = brother_tk.earnings_estimate
                except:
                    pass
            
            eps_1 = None
            eps_2 = None
            
            if estimates is not None and not estimates.empty:
                try:
                    if '0y' in estimates.index: eps_1 = estimates.loc['0y', 'avg']
                    if '+1y' in estimates.index: eps_2 = estimates.loc['+1y', 'avg']
                except:
                    pass

            # 4. NTM Calculation (가중 평균)
            days_remaining = (next_fye_date - today).days
            w = max(0, min(1, (365 - days_remaining) / 365))

            ntm_eps = 0
            status_msg = ""
            
            if eps_1 is not None and eps_2 is not None:
                ntm_eps = (eps_1 * (1 - w)) + (eps_2 * w)
                status_msg = "Standard NTM"
            elif eps_1 is not None:
                ntm_eps = eps_1 
                status_msg = "Partial"
            else:
                fallback = info.get('forwardEps')
                if fallback:
                    ntm_eps = fallback
                    status_msg = "Fallback"
                else:
                    ntm_eps = 0
                    status_msg = "No Data"

            # 5. 지표 산출
            ntm_per = (current_price / ntm_eps) if (ntm_eps and ntm_eps > 0) else 0
            
            implied_net_income = 0
            if market_cap and ntm_eps and current_price:
                # 시총 / PER = 순이익 (또는 시가총액/주가 * EPS)
                implied_net_income = (market_cap / current_price) * ntm_eps

            # 6. 저장
            results.append({
                'Date': today_str,
                'Ticker': symbol,
                'Name': short_name,
                'Sector': sector,
                'Price': current_price,
                'Analyst_Count': analyst_count,
                'Market_Cap': market_cap,        
                'NTM_EPS': ntm_eps,
                'NTM_PER': round(ntm_per, 2),
                'Implied_Net_Income': implied_net_income, 
                'Status': status_msg,
                'Country': 'KOR' if symbol.endswith('.KS') else 'USA'
            })
            
        except Exception:
            continue

    return pd.DataFrame(results)

# ==========================================
# 5. Summary (검증 가능하도록 정확히 합산)
# ==========================================
def calculate_market_summary(df_usa, df_kor):
    # 오늘 날짜
    today_str = datetime.now().strftime("%Y-%m-%d")
    summary_data = []
    
    # 지수 현재가 가져오기 (Current Price 부활)
    sp500_price = get_index_price('^GSPC')
    kospi_price = get_index_price('^KS200')
    
    # [검증 포인트] 여기서 df_usa는 '오늘 수집한 데이터'만 들어있으므로
    # sum()을 하면 정확히 '오늘자 종목들의 합계'가 나옵니다.
    
    # 1. USA Summary
    if not df_usa.empty:
        total_cap = df_usa['Market_Cap'].sum()
        total_income = df_usa['Implied_Net_Income'].sum()
        market_per = total_cap / total_income if total_income > 0 else 0
        
        summary_data.append({
            'Date': today_str,              # A열
            'Region': 'USA',                # B열
            'Index': 'S&P 500',             # C열
            'Current Price': sp500_price,   # D열
            'Total Market Cap': total_cap,  # E열
            'Total Net Income': total_income, # F열 - 사용자님 말씀이 맞습니다.
            'Market PER': round(market_per, 2) # G열
        })

    # 2. KOR Summary
    if not df_kor.empty:
        total_cap = df_kor['Market_Cap'].sum()
        total_income = df_kor['Implied_Net_Income'].sum()
        market_per = total_cap / total_income if total_income > 0 else 0
        
        summary_data.append({
            'Date': today_str, 
            'Region': 'KOR', 
            'Index': 'KOSPI Top 200',
            'Current Price': kospi_price,      # 요청하신 지수 현재가
            'Total Market Cap': total_cap,     # 시가총액 합계
            'Total Net Income': total_income,  # 순이익 합계
            'Market PER': round(market_per, 2)
        })
        
    return pd.DataFrame(summary_data)

# ==========================================
# 6. Sector 분석
# ==========================================
def calculate_sector_summary(df_usa):
    if df_usa.empty: return pd.DataFrame()
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    df_clean = df_usa[df_usa['Sector'] != 'Unknown']
    sector_grp = df_clean.groupby('Sector')[['Market_Cap', 'Implied_Net_Income']].sum().reset_index()
    
    sector_grp['Sector PER'] = sector_grp.apply(
        lambda x: x['Market_Cap'] / x['Implied_Net_Income'] if x['Implied_Net_Income'] > 0 else 0, axis=1
    )
    sector_grp['Date'] = today_str
    sector_grp['Sector PER'] = sector_grp['Sector PER'].round(2)
    
    return sector_grp[['Date', 'Sector', 'Market_Cap', 'Implied_Net_Income', 'Sector PER']]

# ==========================================
# 7. 추세 리포트 (시간 제거 적용)
# ==========================================
def create_trend_report(file_name):
    print("\n📈 [분석] 종목별 추세(EPS Trend) 분석 중...")
    try:
        if not os.path.exists(file_name): return

        df_usa = pd.read_excel(file_name, sheet_name='USA_Stocks')
        df_kor = pd.read_excel(file_name, sheet_name='KOR_Stocks')
        df_combined = pd.concat([df_usa, df_kor], ignore_index=True)
        
        # [중요] 시간 제거 (00:00:00) -> 엑셀 그룹화 필수
        df_combined['Date'] = pd.to_datetime(df_combined['Date']).dt.normalize()
        
        trend_data = []
        unique_tickers = df_combined['Ticker'].unique()
        
        for ticker in unique_tickers:
            history = df_combined[df_combined['Ticker'] == ticker].sort_values('Date')
            if len(history) < 2: continue
            
            history = history.tail(12) 
            start_row = history.iloc[0]
            end_row = history.iloc[-1]
            
            eps_growth = 0
            if start_row['NTM_EPS'] and start_row['NTM_EPS'] != 0:
                eps_growth = ((end_row['NTM_EPS'] - start_row['NTM_EPS']) / abs(start_row['NTM_EPS'])) * 100
            
            price_chg = 0
            if start_row['Price'] and start_row['Price'] != 0:
                price_chg = ((end_row['Price'] - start_row['Price']) / start_row['Price']) * 100
            
            signal = "-"
            if eps_growth > 3 and price_chg < 5: signal = "Chance 🟢"
            elif eps_growth > 10: signal = "Hot 🔥"
            elif eps_growth > 0: signal = "Good 🔺"
            elif eps_growth < 0: signal = "Bad 📉"
            
            trend_data.append({
                'Signal': signal, 'Region': end_row['Country'], 'Ticker': ticker,
                'Name': end_row['Name'], 'Sector': end_row['Sector'],
                'Market_Cap': end_row['Market_Cap'],
                'EPS Growth(%)': round(eps_growth, 2),
                'Price Chg(%)': round(price_chg, 2),
                'Start Date': start_row['Date'].strftime('%Y-%m-%d') if pd.notna(start_row['Date']) else "",
                'End Date': end_row['Date'].strftime('%Y-%m-%d') if pd.notna(end_row['Date']) else ""
            })
            
        df_trend = pd.DataFrame(trend_data)
        if not df_trend.empty:
            df_trend = df_trend.sort_values(by='Market_Cap', ascending=False)
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_trend.to_excel(writer, sheet_name='EPS_Trend', index=False)
                print(" -> [EPS_Trend] 생성 완료 (시간 정보 제거됨).")
    except Exception as e:
        print(f" -> 추세 리포트 실패: {e}")

# ==========================================
# 8. 저장 및 서식 (서식 기능 복구)
# ==========================================
def save_and_format(df_usa, df_kor, df_sum, df_sec):
    print("=== 엑셀 저장 및 서식 적용 중... ===")
    
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        wb.save(FILE_NAME)
        wb.close()
    
    # [수정] Pandas를 사용하여 기존 데이터와 병합 (스키마 변경 대응)
    sheets_to_save = {
        'USA_Stocks': df_usa,
        'KOR_Stocks': df_kor,
        'Summary': df_sum,
        'Sector_Trend': df_sec
    }
    
    final_dfs = {}
    
    try:
        # 기존 파일 구조 읽기
        xls = pd.ExcelFile(FILE_NAME)
        existing_sheets = xls.sheet_names
        
        for sheet_name, new_df in sheets_to_save.items():
            if new_df.empty: continue
            
            if sheet_name in existing_sheets:
                try:
                    # 기존 시트 데이터 읽기
                    old_df = pd.read_excel(xls, sheet_name=sheet_name)
                    # 데이터 병합 (컬럼명 기준 자동 정렬, 새 컬럼 추가됨)
                    combined = pd.concat([old_df, new_df], ignore_index=True)
                    final_dfs[sheet_name] = combined
                except Exception:
                    final_dfs[sheet_name] = new_df
            else:
                final_dfs[sheet_name] = new_df
                
    except Exception as e:
        print(f"⚠️ 기존 파일 읽기 실패 (새 파일로 시작): {e}")
        final_dfs = sheets_to_save

    # 저장 (덮어쓰기 모드)
    with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for sheet_name, df in final_dfs.items():
            # NaN을 빈 문자열이나 0으로 처리하지 않고 그대로 둠 (Excel에서 처리)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 서식 적용
    wb = load_workbook(FILE_NAME)
    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
    
    for sheet_name in wb.sheetnames:
        if sheet_name not in ['USA_Stocks', 'KOR_Stocks', 'Summary', 'Sector_Trend']: continue
        
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
        # 헤더 읽기
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = headers[cell.column - 1]
                
                # A열 날짜 서식
                if cell.column == 1: 
                    cell.number_format = 'yyyy-mm-dd'
                
                # 숫자 서식
                if isinstance(cell.value, (int, float)):
                    if cell.value > 1000000000 and header == 'Market_Cap':
                         # 조 단위 / T 단위 구분
                        if sheet_name == 'USA_Stocks':
                            cell.number_format = '#,##0,,," T"'
                        elif sheet_name == 'KOR_Stocks':
                            cell.number_format = '#,##0,,,,"조"'
                        else:
                            cell.number_format = '#,##0'
                    elif header == 'Analyst_Count':
                        cell.number_format = '0' # 정수
                    else:
                        cell.number_format = '0.00'

for sheet_name in wb.sheetnames:
        # L열과 M열 교체는 USA_Stocks와 KOR_Stocks 시트에만 적용합니다.
        if sheet_name in ['USA_Stocks', 'KOR_Stocks']:
            ws = wb[sheet_name]
            # 최소 13열이 있어야 교체 가능 (A=1, L=12, M=13)
            if ws.max_column >= 13:
                # 교체할 헤더 값 스왑
                h_l = ws.cell(row=1, column=12).value
                h_m = ws.cell(row=1, column=13).value
                ws.cell(row=1, column=12).value = h_m
                ws.cell(row=1, column=13).value = h_l

                # 각 행의 값 스왑
                for r in range(2, ws.max_row + 1):
                    v_l = ws.cell(row=r, column=12).value
                    v_m = ws.cell(row=r, column=13).value
                    ws.cell(row=r, column=12).value = v_m
                    ws.cell(row=r, column=13).value = v_l

        wb.save(FILE_NAME)
        print("=== 모든 작업 완료 (Analyst_Count 반영 및 L/M 열 스왑 완료) ===")

# ==========================================
# 9. 메인 실행 (안전장치 + 검증용 로직)
# ==========================================
if __name__ == "__main__":
    print(f"=== [V21 Final] 주식 데이터 자동화 시작 ===")
    
    # 1. 수집
    usa_tickers = get_sp500_tickers()
    kor_tickers = get_kospi200_cleaned_tickers()
    full_tickers = usa_tickers + kor_tickers
    
    if full_tickers:
        # 2. 분석 (오늘치 데이터만 수집)
        df_results = analyze_ntm_data_robust(full_tickers)
        
        if not df_results.empty:
            df_usa = df_results[df_results['Country'] == 'USA']
            df_kor = df_results[df_results['Country'] == 'KOR']
            
            # 3. 요약 (df_usa는 오늘 데이터만 있으므로 sum()은 오늘치 합계와 동일함 -> 검증 완료)
            df_sum = calculate_market_summary(df_usa, df_kor)
            
            # 4. 섹터
            df_sec = calculate_sector_summary(df_usa)
            
            # 5. 저장
            save_and_format(df_usa, df_kor, df_sum, df_sec)
            
            # 6. 추세 리포트
            create_trend_report(FILE_NAME)

            # ---------------------------------------------------------
            # 7. 구글 시트 자동 업로드 (새로 추가된 클라우드 동기화)
            # ---------------------------------------------------------
            print("\n☁️ 구글 시트 업로드를 시작합니다...")
            try:
                scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
                creds = ServiceAccountCredentials.from_json_keyfile_name('stock-key.json', scope)
                client = gspread.authorize(creds)

                # 엑셀의 여러 시트를 클라우드로 올립니다. (Summary 포함 모든 주요 시트)
                sheet_names = ['Summary', 'USA_Stocks', 'KOR_Stocks', 'Sector_Trend', 'EPS_Trend']
                upload_excel_sheets_to_gspread(client, 'Stock Data', FILE_NAME, sheet_names)
                
                print(f"🚀 구글 시트 'Stock Data' 업데이트 성공!")
                
            except Exception as upload_error:
                print(f"❌ 구글 시트 업로드 중 오류 발생: {upload_error}")
            # ---------------------------------------------------------

            print(f"\n✅ [최종 성공] 모든 분석 및 클라우드 동기화가 완료되었습니다!")
            
        else:
            print("\n❌ [오류] 분석된 데이터가 없습니다.")
    else:
        print("\n❌ [오류] 종목 리스트를 가져오지 못했습니다.")